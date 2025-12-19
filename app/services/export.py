import csv
import io
from datetime import date
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from app.models.work_week import WorkWeek
from app.models.work_item import WorkItem, TaskType, TaskStatus


def get_filtered_items(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    task_type: Optional[str] = None,
    status: Optional[str] = None
) -> List[dict]:
    """Get work items with filters applied."""
    query = db.query(WorkItem).join(WorkWeek)
    
    if start_date:
        query = query.filter(WorkWeek.week_start >= start_date)
    if end_date:
        query = query.filter(WorkWeek.week_end <= end_date)
    if task_type:
        query = query.filter(WorkItem.type == task_type)
    if status:
        query = query.filter(WorkItem.status == status)
    
    items = query.order_by(WorkWeek.week_start.desc(), WorkItem.created_at).all()
    
    return [{
        "Week Start": item.work_week.week_start.strftime("%Y-%m-%d"),
        "Week End": item.work_week.week_end.strftime("%Y-%m-%d"),
        "Title": item.title,
        "Type": item.type,
        "Status": item.status,
        "Start Date": item.start_date.strftime("%Y-%m-%d") if item.start_date else "",
        "End Date": item.end_date.strftime("%Y-%m-%d") if item.end_date else "",
        "Assigned Points": item.assigned_points,
        "Completion Points": item.completion_points or "",
        "Planned Work": item.planned_work or "",
        "Actual Work": item.actual_work or "",
        "Next Week Plan": item.next_week_plan or ""
    } for item in items]


def export_to_csv(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    task_type: Optional[str] = None,
    status: Optional[str] = None
) -> str:
    """Export filtered items to CSV string."""
    items = get_filtered_items(db, start_date, end_date, task_type, status)
    
    if not items:
        return "No data found"
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=items[0].keys())
    writer.writeheader()
    writer.writerows(items)
    
    return output.getvalue()


def export_to_excel(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    task_type: Optional[str] = None,
    status: Optional[str] = None
) -> bytes:
    """Export filtered items to Excel bytes."""
    items = get_filtered_items(db, start_date, end_date, task_type, status)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Items"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    
    if items:
        headers = list(items[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row, item in enumerate(items, 2):
            for col, value in enumerate(item.values(), 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Work Tracker Export Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A3"] = f"Total Items: {len(items)}"
    ws_summary["A4"] = f"Export Date: {date.today().strftime('%Y-%m-%d')}"
    
    if items:
        total_points = sum(item["Assigned Points"] for item in items)
        ws_summary["A5"] = f"Total Points: {total_points}"
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
